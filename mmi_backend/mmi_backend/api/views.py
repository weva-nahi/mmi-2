"""
Views DRF — MMI Platform
Tous les ViewSets dans une seule application.
"""
from rest_framework import viewsets, generics, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.conf import settings
from django.utils.encoding import force_str, force_bytes
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.auth.tokens import default_token_generator
import logging

from .models import (
    User, Role, UserRole,
    Actualite, DocumentPublic, ProjetBanque,
    TypeDemande, Demande, EtapeTraitement, PieceRequise, PieceJointe, Notification,
    Autorisation, DocumentGenere,
    VisiteLieux, ComiteBP, DistanceBoulangerie,
    RenouvellementDetail, ExtensionDetail, UsineEauDetail, UniteIndustrielleDetail,
)
from .serializers import (
    CustomTokenObtainPairSerializer, InscriptionSerializer,
    UserListSerializer, UserDetailSerializer, UserCreateSerializer,
    RoleSerializer, UserRoleSerializer,
    ActualiteSerializer, DocumentPublicSerializer, ProjetBanqueSerializer,
    TypeDemandeSerializer, PieceRequiseSerializer, PieceJointeSerializer,
    DemandeListSerializer, DemandeDetailSerializer,
    EtapeTraitementSerializer, TransitionStatutSerializer,
    NotificationSerializer,
    AutorisationSerializer, AutorisationGeoJSONSerializer, DocumentGenereSerializer,
    VisiteLieuxSerializer, ComiteBPSerializer, DistanceBoulangerieSerializer,
    RenouvellementDetailSerializer, ExtensionDetailSerializer,
    UsineEauDetailSerializer, UniteIndustrielleDetailSerializer,
)
from .permissions import (
    IsSuperAdmin, IsDemandeur, IsSecretariatCentral, IsSecretaireGeneral,
    IsMinistre, IsDGI, IsDGIDirecteur, IsDDPI, IsMMISignataire,
    IsAgentInstitutionnel, IsOwnerOrAgent,
    IsDDPIDirecteur, IsDDPIChefBP, IsDDPIChefUsines,
    IsSecretaireComiteBP, IsDGISecretariat,
)

logger = logging.getLogger('api')


# ─────────────────────────────────────────────────────────────
# AUTHENTIFICATION
# ─────────────────────────────────────────────────────────────

class LoginView(TokenObtainPairView):
    """POST /api/auth/login/ — connexion via identifiant_unique + password."""
    permission_classes = [AllowAny]
    serializer_class   = CustomTokenObtainPairSerializer


class InscriptionView(generics.CreateAPIView):
    """POST /api/auth/register/ — auto-inscription du demandeur."""
    permission_classes = [AllowAny]
    serializer_class   = InscriptionSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
            'message': 'Compte créé avec succès. Votre identifiant a été envoyé par email.',
            'identifiant_unique': user.identifiant_unique,
            'email': user.email,
        }, status=status.HTTP_201_CREATED)


# ─────────────────────────────────────────────────────────────
# UTILISATEURS & RÔLES (Super Admin)
# ─────────────────────────────────────────────────────────────

class UserViewSet(viewsets.ModelViewSet):
    """
    CRUD utilisateurs — réservé Super Admin.
    GET /api/admin/users/
    POST /api/admin/users/
    """
    queryset           = User.objects.all().order_by('-created_at')
    permission_classes = [IsSuperAdmin]
    filter_backends    = [DjangoFilterBackend, filters.SearchFilter]
    search_fields      = ['nom', 'prenom', 'email', 'nom_entreprise', 'identifiant_unique']
    filterset_fields   = ['is_active', 'is_super_admin']

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action in ['retrieve', 'update', 'partial_update']:
            return UserDetailSerializer
        return UserListSerializer

    @action(detail=True, methods=['post'], url_path='attribuer-role')
    def attribuer_role(self, request, pk=None):
        """POST /api/admin/users/{id}/attribuer-role/ — attribuer un rôle (remplace tous les anciens)."""
        user = self.get_object()
        role_code = request.data.get('role_code')
        if not role_code:
            return Response({'error': 'role_code requis.'}, status=400)
        try:
            role = Role.objects.get(code=role_code)
        except Role.DoesNotExist:
            return Response({'error': f'Rôle {role_code} introuvable.'}, status=400)

        # Supprimer TOUS les anciens rôles de l'utilisateur
        UserRole.objects.filter(user=user).delete()

        # Créer le nouveau rôle unique
        UserRole.objects.create(
            user=user, role=role,
            actif=True, assigned_by=request.user
        )
        logger.info(f"Rôle de {user.email} changé → {role_code} par {request.user.email}")
        return Response({
            'message': f'Rôle de {user.nom_complet} changé en {role.nom}.',
            'role': role_code,
        })

    @action(detail=True, methods=['post'], url_path='activer')
    def activer(self, request, pk=None):
        """Activer ou suspendre un compte."""
        user = self.get_object()

        # Bloquer la désactivation de son propre compte
        if user.pk == request.user.pk:
            return Response(
                {'detail': 'Vous ne pouvez pas désactiver votre propre compte.'},
                status=400
            )

        # Bloquer la désactivation d'un super admin
        if user.is_super_admin and not user.is_active == False:
            return Response(
                {'detail': 'Le compte Super Admin ne peut pas être suspendu.'},
                status=400
            )

        user.is_active = not user.is_active
        user.save()
        action_label = "activé" if user.is_active else "suspendu"
        return Response({'message': f"Compte {action_label}.", 'is_active': user.is_active})


class RoleViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = Role.objects.all()
    serializer_class   = RoleSerializer
    permission_classes = [IsAgentInstitutionnel]


# ─────────────────────────────────────────────────────────────
# PORTAIL PUBLIC
# ─────────────────────────────────────────────────────────────

class ActualiteViewSet(viewsets.ModelViewSet):
    serializer_class = ActualiteSerializer
    filter_backends  = [filters.SearchFilter, DjangoFilterBackend]
    search_fields    = ['titre', 'contenu', 'slug']
    filterset_fields = ['publie']

    def get_queryset(self):
        if self.request.user.is_authenticated and self.request.user.is_super_admin:
            return Actualite.objects.all()
        return Actualite.objects.filter(publie=True)

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'by_slug']:
            return [AllowAny()]
        return [IsSuperAdmin()]

    def perform_create(self, serializer):
        serializer.save(auteur=self.request.user)

    @action(detail=False, methods=['get'], url_path='by-slug/(?P<slug>[^/.]+)',
            permission_classes=[AllowAny])
    def by_slug(self, request, slug=None):
        """GET /api/public/actualites/by-slug/{slug}/ — récupère un article par slug."""
        try:
            qs = self.get_queryset()
            article = qs.get(slug=slug)
            return Response(ActualiteSerializer(article).data)
        except Actualite.DoesNotExist:
            return Response({'detail': 'Article introuvable.'}, status=404)


class DocumentPublicViewSet(viewsets.ModelViewSet):
    serializer_class = DocumentPublicSerializer
    filter_backends  = [DjangoFilterBackend]
    filterset_fields = ['categorie', 'langue', 'publie']

    def get_queryset(self):
        if self.request.user.is_authenticated and self.request.user.is_super_admin:
            return DocumentPublic.objects.all()
        return DocumentPublic.objects.filter(publie=True)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsSuperAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ProjetBanqueViewSet(viewsets.ModelViewSet):
    serializer_class = ProjetBanqueSerializer
    filter_backends  = [DjangoFilterBackend]
    filterset_fields = ['statut', 'secteur', 'publie']

    def get_queryset(self):
        if self.request.user.is_authenticated and self.request.user.is_super_admin:
            return ProjetBanque.objects.all()
        return ProjetBanque.objects.filter(publie=True)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsSuperAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


# ─────────────────────────────────────────────────────────────
# DEMANDES
# ─────────────────────────────────────────────────────────────

class TypeDemandeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = TypeDemande.objects.filter(actif=True).order_by('id')
    serializer_class   = TypeDemandeSerializer
    permission_classes = [AllowAny]


class ImprimerDossierView(generics.GenericAPIView):
    """
    GET /api/demandes/{id}/imprimer/
    Retourne toutes les données du dossier pour impression papier (DGI Secrétariat).
    Déclenche aussi la transition DGI_SEC_IMPRESSION.
    """
    permission_classes = [IsAgentInstitutionnel]

    def get(self, request, pk=None):
        from api.models import Demande, EtapeTraitement
        try:
            demande = Demande.objects.get(pk=pk)
        except Demande.DoesNotExist:
            return Response({'detail': 'Dossier introuvable.'}, status=404)

        # Enregistrer l'étape d'impression si pas déjà fait
        deja_imprime = EtapeTraitement.objects.filter(
            demande=demande, etape_code='DGI_SEC_IMPRESSION'
        ).exists()

        if not deja_imprime:
            statut_avant = demande.statut
            demande.statut = 'PRET_IMPRESSION_DGI'
            demande.save()
            EtapeTraitement.objects.create(
                demande=demande,
                acteur=request.user,
                etape_code='DGI_SEC_IMPRESSION',
                statut_avant=statut_avant,
                statut_apres='PRET_IMPRESSION_DGI',
                action='Dossier mis à disposition pour impression — circuit papier',
            )

        # Construire les données complètes pour impression
        from api.serializers import DemandeDetailSerializer
        data = DemandeDetailSerializer(demande, context={'request': request}).data
        return Response({
            'dossier':     data,
            'imprimable':  True,
            'message':     'Dossier prêt pour impression.',
        })


class PieceRequiseViewSet(viewsets.ModelViewSet):
    """CRUD pièces requises — lecture publique, écriture Super Admin."""
    serializer_class = PieceRequiseSerializer
    filter_backends  = [DjangoFilterBackend]
    filterset_fields = ['type_demande', 'obligatoire']

    def get_queryset(self):
        qs = PieceRequise.objects.all().order_by('type_demande', 'ordre')
        type_demande = self.request.query_params.get('type_demande')
        if type_demande:
            qs = qs.filter(type_demande_id=type_demande)
        return qs

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsSuperAdmin()]



# ─────────────────────────────────────────────────────────────
# HELPER : Notifications automatiques par étape
# ─────────────────────────────────────────────────────────────

def _notifier_prochaine_etape(demande, etape_code, acteur_courant):
    """
    Envoie une notification email à l'acteur suivant dans le circuit.
    Appelé après chaque transition dans transmettre().
    """
    from api.models import Notification
    from django.core.mail import send_mail
    from django.conf import settings

    # Mapping etape_code → (rôle suivant, message)
    NOTIFICATIONS = {
        'SC_RECEPTION': {
            'role':    'SEC_CENTRAL',
            'titre':   'Accusé de réception enregistré',
            'message': f"Le dossier {demande.numero_ref} a été réceptionné. Prêt pour transmission au SG.",
        },
        'SC_TRANSMISSION_SG': {
            'role':    'SEC_GENERAL',
            'titre':   'Nouveau dossier à traiter — Secrétariat Général',
            'message': f"Le dossier {demande.numero_ref} ({demande.type_demande.libelle}) vous a été transmis par le Secrétariat Central.",
        },
        'DGI_SEC_IMPRESSION': {
            'role':    'DGI_SECRETARIAT',
            'titre':   'Dossier disponible — Impression requise',
            'message': f"Le dossier {demande.numero_ref} est disponible pour impression. Merci de procéder à l'impression pour le circuit papier.",
        },
        'SG_TRANSMISSION_MIN': {
            'role':    'MINISTRE',
            'titre':   'Dossier soumis à votre appréciation — Ministre',
            'message': f"Le Secrétaire Général vous a transmis le dossier {demande.numero_ref} ({demande.type_demande.libelle}) pour lecture et transmission.",
        },
        'SG_TRANSMISSION_DGI': {
            'role':    'DGI_DIRECTEUR',
            'titre':   'Nouveau dossier en instruction — DGI',
            'message': f"Le dossier {demande.numero_ref} vous a été transmis directement par le SG pour instruction technique.",
        },
        'MIN_TRANSMISSION_DGI': {
            'role':    'DGI_DIRECTEUR',
            'titre':   'Nouveau dossier en instruction — DGI',
            'message': f"Le dossier {demande.numero_ref} vous a été transmis par le Ministre pour instruction technique.",
        },
        'DGI_TRANSMISSION_DDPI': {
            'role':    'DDPI_DIRECTEUR',
            'titre':   'Dossier transmis à la DDPI — Action requise',
            'message': f"La DGI vous a transmis le dossier {demande.numero_ref} ({demande.type_demande.libelle}) pour traitement.",
        },
        'DDPI_CHEF_BP': {
            'role':    'DDPI_CHEF_BP',
            'titre':   'Dossier BP assigné — Chef Service Boulangeries',
            'message': f"Le dossier BP {demande.numero_ref} vous a été orienté pour traitement.",
        },
        'DDPI_CHEF_USINES': {
            'role':    'DDPI_CHEF_USINES',
            'titre':   'Dossier Usine assigné — Chef Service Usines',
            'message': f"Le dossier {demande.numero_ref} ({demande.type_demande.libelle}) vous a été orienté pour traitement.",
        },
        'DDPI_COMMISSION_BP': {
            'role':    'SEC_COMITE_BP',
            'titre':   'Dossier soumis en commission BP — PV requis',
            'message': f"Le dossier {demande.numero_ref} est soumis en commission BP. Merci d'enregistrer le résultat et de déposer le PV.",
        },
        'DDPI_PV_COMITE_BP': {
            'role':    'DDPI_CHEF_BP',
            'titre':   'PV comité BP déposé — Quittance attendue',
            'message': f"Le PV du comité BP pour le dossier {demande.numero_ref} a été déposé. En attente de la quittance de paiement.",
        },
        'DDPI_ACCORD_PRINCIPE': {
            'role':    'DGI_DIRECTEUR',
            'titre':   'Accord de principe accordé — Signature DGI requise',
            'message': f"La DDPI a accordé l'accord de principe pour {demande.numero_ref}. Signature DGI requise.",
        },
        'DGI_SIGNATURE': {
            'role':    'MMI_SIGNATAIRE',
            'titre':   'Arrêté prêt pour signature — Ministre / Signataire MMI',
            'message': f"Le dossier {demande.numero_ref} est prêt pour la signature officielle du Ministre.",
        },
        'DDPI_INCOMPLET': {
            'role':    'DEMANDEUR',
            'titre':   'Action requise — Dossier incomplet',
            'message': f"Votre dossier {demande.numero_ref} présente des éléments manquants. Connectez-vous pour consulter le motif.",
        },
        'DDPI_REJET': {
            'role':    'DEMANDEUR',
            'titre':   'Décision sur votre dossier',
            'message': f"Votre dossier {demande.numero_ref} a fait l'objet d'une décision. Connectez-vous pour en consulter le détail.",
        },
    }

    notif_config = NOTIFICATIONS.get(etape_code)
    if not notif_config:
        return

    role_cible = notif_config['role']
    titre      = notif_config['titre']
    message    = notif_config['message']

    # Trouver les utilisateurs du rôle cible
    from api.models import UserRole, Notification as Notif, User
    if role_cible == 'DEMANDEUR':
        destinataires = [demande.demandeur]
    else:
        user_ids = UserRole.objects.filter(
            role__code=role_cible, actif=True
        ).values_list('user_id', flat=True)
        destinataires = User.objects.filter(id__in=user_ids, is_active=True)

    for dest in destinataires:
        # Notification en base
        Notif.objects.create(
            destinataire=dest,
            demande=demande,
            type_notif='ACTION_REQUISE',
            titre=titre,
            message=message,
        )
        # Email
        if dest.email:
            try:
                send_mail(
                    subject=f"[MMI] {titre}",
                    message=f"Bonjour {dest.nom_complet},\n\n{message}\n\nConnectez-vous sur : https://plateforme.mmi.gov.mr\n\nCordialement,\nPlateforme MMIAPP",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[dest.email],
                    fail_silently=True,
                )
            except Exception:
                pass


class DemandeViewSet(viewsets.ModelViewSet):
    """
    CRUD demandes avec gestion des droits par rôle.
    Le demandeur ne voit que ses propres dossiers.
    Les agents voient tous les dossiers selon leur rôle.
    """
    filter_backends  = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['statut', 'type_demande__code', 'wilaya']
    search_fields    = ['numero_ref', 'raison_sociale', 'activite']
    ordering_fields  = ['date_soumission', 'date_maj', 'statut']
    ordering         = ['-date_soumission']

    def get_queryset(self):
        user = self.request.user
        if user.is_super_admin:
            return Demande.objects.select_related('demandeur', 'type_demande').all()
        if user.has_role('DEMANDEUR'):
            return Demande.objects.filter(demandeur=user)
        # Tous les agents voient toutes les demandes
        return Demande.objects.select_related('demandeur', 'type_demande').all()

    def get_serializer_class(self):
        if self.action == 'list':
            return DemandeListSerializer
        return DemandeDetailSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [IsDemandeur()]
        return [IsAuthenticated()]

    @action(detail=True, methods=['post'], url_path='transmettre',
            permission_classes=[IsAgentInstitutionnel])
    def transmettre(self, request, pk=None):
        """
        POST /api/demandes/{id}/transmettre/
        Effectue une transition de statut avec enregistrement dans le journal.
        Supporte pièce jointe via multipart/form-data.
        """
        demande = self.get_object()
        serializer = TransitionStatutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data
        statut_avant = demande.statut

        # Mapping etape_code → nouveau statut
        TRANSITIONS = {
            # Secrétariat Central
            'SC_RECEPTION':           'EN_RECEPTION',
            'SC_TRANSMISSION_SG':     'TRANSMISE_SG',
            # Secrétariat DGI — impression
            'DGI_SEC_IMPRESSION':     'PRET_IMPRESSION_DGI',
            # Secrétaire Général
            'SG_LECTURE':             'EN_LECTURE_SG',
            'SG_TRANSMISSION_MIN':    'TRANSMISE_MINISTRE',
            'SG_TRANSMISSION_DGI':    'EN_INSTRUCTION_DGI',
            # Ministre
            'MIN_LECTURE':            'EN_LECTURE_MINISTRE',
            'MIN_TRANSMISSION_DGI':   'EN_INSTRUCTION_DGI',
            # DGI
            'DGI_INSTRUCTION':        'EN_INSTRUCTION_DGI',
            'DGI_TRANSMISSION_DDPI':  'EN_TRAITEMENT_DDPI',
            'DGI_SIGNATURE':          'SIGNATURE_DGI',
            # DDPI Directeur — orientation
            'DDPI_VERIFICATION':      'EN_TRAITEMENT_DDPI',
            'DDPI_INCOMPLET':         'DOSSIER_INCOMPLET',
            'DDPI_CHEF_BP':           'EN_TRAITEMENT_DDPI_BP',
            'DDPI_CHEF_USINES':       'EN_TRAITEMENT_DDPI_US',
            # DDPI Chefs de service
            'DDPI_VISITE':            'VISITE_PROGRAMMEE',
            'DDPI_PV_VISITE':         'EN_TRAITEMENT_DDPI',
            'DDPI_COMMISSION_BP':     'EN_COMMISSION_BP',
            'DDPI_PV_COMITE_BP':      'PV_COMITE_DEPOSE',
            'DDPI_QUITTANCE_BP':      'ATTENTE_QUITTANCE',
            'DDPI_ACCORD_PRINCIPE':   'ACCORD_PRINCIPE',
            'DDPI_ARRETE_REDACTION':  'ARRETE_EN_COURS',
            'DDPI_REJET':             'REJETE',
            # MMI / Ministre
            'MMI_SIGNATURE':          'SIGNATURE_MMI',
            'SYSTEME_VALIDATION':     'VALIDE',
        }

        nouveau_statut = TRANSITIONS.get(data['etape_code'])
        if nouveau_statut:
            demande.statut = nouveau_statut
            demande.save()

        etape = EtapeTraitement(
            demande=demande,
            acteur=request.user,
            etape_code=data['etape_code'],
            statut_avant=statut_avant,
            statut_apres=demande.statut,
            action=data['action'],
            commentaire=data.get('commentaire', ''),
        )
        # Pièce jointe optionnelle
        if 'piece_jointe' in request.FILES:
            etape.piece_jointe = request.FILES['piece_jointe']
        etape.save()

        logger.info(f"Demande {demande.numero_ref} : {statut_avant} → {demande.statut} par {request.user.nom_complet}")

        return Response({
            'message': 'Transition effectuée.',
            'numero_ref': demande.numero_ref,
            'statut_avant': statut_avant,
            'statut_apres': demande.statut,
        })

    @action(detail=True, methods=['get'], url_path='historique')
    def historique(self, request, pk=None):
        """GET /api/demandes/{id}/historique/ — journal d'audit complet."""
        demande = self.get_object()
        etapes  = demande.etapes.all()
        return Response(EtapeTraitementSerializer(etapes, many=True).data)

    @action(detail=True, methods=['post'], url_path='valider-final',
            permission_classes=[IsAgentInstitutionnel])
    def valider_final(self, request, pk=None):
        """
        POST /api/demandes/{id}/valider-final/
        Signature MMI → génère l'autorisation officielle + valide la demande.
        """
        demande = self.get_object()

        # Créer l'autorisation officielle
        from datetime import date, timedelta
        from api.models import Autorisation, DocumentGenere

        # Numéro autorisation : AUTO-{YEAR}-{TYPE}-{ID:04d}
        annee = date.today().year
        numero = f"AUTO-{annee}-{demande.type_demande.code}-{demande.id:04d}"

        # Durée fixe : 6 mois (182 jours) pour tous les types
        # Pour un RENOUVELLEMENT : la nouvelle expiration part de l'ancienne date d'expiration
        # (pas d'aujourd'hui, pour ne pas pénaliser ceux qui renouvellent en avance)
        DUREE_JOURS = 182  # 6 mois

        if demande.type_demande.code == 'RENOUVELLEMENT' and demande.autorisation_parent:
            # Renouvellement : date_exp = ancienne_exp + 6 mois
            ancienne_exp = demande.autorisation_parent.date_expiration
            base = ancienne_exp if ancienne_exp and ancienne_exp > date.today() else date.today()
            date_exp = base + timedelta(days=DUREE_JOURS)
            # Marquer l'ancienne autorisation comme expirée/renouvelée
            demande.autorisation_parent.statut = 'expiree'
            demande.autorisation_parent.save()

        elif demande.type_demande.code == 'EXTENSION' and demande.autorisation_parent:
            # Extension : hérite de la date d'expiration de l'autorisation d'origine
            date_exp = demande.autorisation_parent.date_expiration

        elif demande.type_demande.code == 'EXTENSION':
            # Extension sans autorisation parente — pas de date propre
            date_exp = None

        else:
            # BP, USINE_EAU, UNITE : 6 mois à partir d'aujourd'hui
            date_exp = date.today() + timedelta(days=DUREE_JOURS)

        # Créer ou récupérer l'autorisation
        auto, created = Autorisation.objects.get_or_create(
            demande=demande,
            defaults={
                'numero_auto':     numero,
                'type':            demande.type_demande.code,
                'statut':          'active',
                'wilaya':          demande.wilaya or '',
                'adresse':         demande.adresse or '',
                'latitude':        demande.latitude,    # GPS auto depuis la demande
                'longitude':       demande.longitude,
                'date_delivrance': date.today(),
                'date_expiration': date_exp,
            }
        )
        # Si autorisation déjà existante, mettre à jour le GPS
        if not created and demande.latitude and not auto.latitude:
            auto.latitude  = demande.latitude
            auto.longitude = demande.longitude
            auto.save()

        # Changer statut demande → VALIDE
        statut_avant = demande.statut
        demande.statut = 'VALIDE'
        demande.save()

        # Enregistrer l'étape MMI_SIGNATURE
        EtapeTraitement.objects.create(
            demande=demande,
            acteur=request.user,
            etape_code='MMI_SIGNATURE',
            statut_avant=statut_avant,
            statut_apres='VALIDE',
            action="Signature officielle de l'arrêté par le MMI — Autorisation délivrée",
            commentaire=request.data.get('commentaire', ''),
        )

        # Aussi enregistrer l'étape système VALIDATION
        EtapeTraitement.objects.create(
            demande=demande,
            acteur=request.user,
            etape_code='SYSTEME_VALIDATION',
            statut_avant='VALIDE',
            statut_apres='VALIDE',
            action="Autorisation officielle générée et disponible au téléchargement",
            commentaire='',
        )

        # Notification au demandeur
        from api.models import Notification
        Notification.objects.create(
            destinataire=demande.demandeur,
            demande=demande,
            type_notif='VALIDATION',
            titre='✅ Votre autorisation est prête !',
            message=f"Félicitations ! Votre demande {demande.numero_ref} a été approuvée. "                    f"Votre autorisation N° {numero} est disponible sur votre espace."                    + (f" Date d'expiration : {date_exp.strftime('%d/%m/%Y')}." if date_exp else ""),
        )

        # Générer le PDF d'autorisation
        try:
            import sys, os
            sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            from generate_autorisation import generer_autorisation, CHAMPS_PAR_TYPE
            from django.conf import settings as dj_settings
            from api.models import ConfigPlateforme

            # Récupérer les données dynamiques de la demande
            dem = demande
            config_type = CHAMPS_PAR_TYPE.get(dem.type_demande.code, CHAMPS_PAR_TYPE['BP'])

            # Construire le data dict dynamiquement
            pdf_data = {
                'type_demande':           dem.type_demande.code,
                'numero_auto':            numero,
                'numero_ref':             dem.numero_ref,
                'identifiant_demandeur':  dem.demandeur.identifiant_unique or dem.numero_ref,
                'nom_responsable':        dem.demandeur.nom_complet,
                'nif':                    getattr(dem.demandeur, 'nif', '') or '',
                'telephone':              getattr(dem.demandeur, 'telephone', '') or '',
                'email':                  dem.demandeur.email,
                'adresse':                getattr(dem.demandeur, 'adresse_siege', '') or dem.adresse or '',
                'forme_juridique':        getattr(dem.demandeur, 'forme_juridique', '') or '',
                'nom_etablissement':      dem.raison_sociale or dem.demandeur.nom_entreprise or '',
                'type_activite':          dem.activite or '',
                'wilaya':                 dem.wilaya or '',
                'adresse_local':          dem.adresse or '',
                'description_activite':   dem.activite or '',
                'capital_social':         '',
                'emplois_directs':        '',
                'ancien_numero':          '',
                'date_delivrance':        date.today().strftime('%d/%m/%Y'),
                'date_expiration':        date_exp.strftime('%d/%m/%Y') if date_exp else '',
                'signataire_nom':         ConfigPlateforme.get('ministre_nom', 'Le Ministre'),
                'signataire_titre':       ConfigPlateforme.get('ministre_titre', "Ministre des Mines et de l'Industrie"),
                'plateforme_url':         ConfigPlateforme.get('plateforme_url', 'https://plateforme.mmi.gov.mr'),
            }

            # Données spécifiques selon type
            if dem.type_demande.code == 'UNITE' and hasattr(dem, 'unite_detail'):
                ud = dem.unite_detail
                pdf_data['capital_social'] = str(ud.capital_social or '')
                pdf_data['emplois_directs'] = str(ud.emplois_directs or '')
                pdf_data['description_activite'] = ud.description_unite or dem.activite or ''
            elif dem.type_demande.code == 'RENOUVELLEMENT' and hasattr(dem, 'renouvellement_detail'):
                rd = dem.renouvellement_detail
                pdf_data['ancien_numero'] = getattr(rd, 'numero_autorisation_original', '') or ''
            elif dem.type_demande.code == 'EXTENSION' and hasattr(dem, 'extension_detail'):
                ed = dem.extension_detail
                pdf_data['ancien_numero'] = getattr(ed, 'numero_autorisation_original', '') or ''
                pdf_data['description_extension'] = getattr(ed, 'description_extension', '') or ''

            # Chemin logo
            logo_path = os.path.join(dj_settings.BASE_DIR, '..', '..', 'mmi_frontend',
                                     'mmi_frontend', 'public', 'images', 'logo_mmi.png')

            # Générer le PDF
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                pdf_path = tmp.name

            generer_autorisation(pdf_data, pdf_path, logo_path if os.path.exists(logo_path) else None)

            # Créer le DocumentGenere lié à la demande
            from django.core.files import File
            # Type de document selon le type de demande
            TYPE_DOC_MAP = {
                'BP':             'arrete_bp',
                'USINE_EAU':      'arrete_conjoint',   # MMI + MCIAT
                'UNITE':          'accord_principe',    # Accord de principe uniquement
                'RENOUVELLEMENT': 'certificat',
                'EXTENSION':      'attestation',
            }
            type_doc_final = TYPE_DOC_MAP.get(demande.type_demande.code, 'attestation')

            doc = DocumentGenere(
                demande=demande,
                type_doc=type_doc_final,
                signataire=request.user,
                signataire_role='MMI_SIGNATAIRE',
                signe=True,
                date_signature=timezone.now(),
            )
            with open(pdf_path, 'rb') as f:
                doc.fichier_pdf.save(f'autorisation_{numero}.pdf', File(f), save=True)
            os.unlink(pdf_path)

        except Exception as e:
            import logging
            logging.getLogger('api').warning(f"Erreur generation PDF: {e}")

        return Response({
            'message':       'Demande validée — Autorisation délivrée.',
            'numero_auto':   numero,
            'date_exp':      str(date_exp) if date_exp else None,
            'autorisation_id': auto.id,
        }, status=201)

    @action(detail=True, methods=['post'], url_path='upload-piece',
            permission_classes=[IsDemandeur])
    def upload_piece(self, request, pk=None):
        """POST /api/demandes/{id}/upload-piece/ — upload d'une pièce jointe."""
        demande = self.get_object()
        if demande.demandeur != request.user:
            return Response({'error': 'Accès refusé.'}, status=403)

    @action(detail=True, methods=['post'], url_path='upload-piece-agent',
            permission_classes=[IsAgentInstitutionnel])
    def upload_piece_agent(self, request, pk=None):
        """POST /api/demandes/{id}/upload-piece-agent/ — agent joint un fichier à une étape."""
        demande  = self.get_object()
        fichier  = request.FILES.get('fichier')
        etape_id = request.data.get('etape_id')
        if not fichier:
            return Response({'detail': 'Aucun fichier fourni.'}, status=400)
        # Attacher à une étape existante si fourni
        if etape_id:
            try:
                etape = EtapeTraitement.objects.get(pk=etape_id, demande=demande)
                etape.piece_jointe = fichier
                etape.save()
                return Response({'message': 'Pièce jointe à l\'étape.', 'etape_id': etape.id})
            except EtapeTraitement.DoesNotExist:
                return Response({'detail': 'Étape introuvable.'}, status=404)
        # Sinon attacher à la demande directement
        from api.models import PieceJointe
        pj = PieceJointe.objects.create(
            demande=demande,
            nom_original=fichier.name,
            fichier=fichier,
            taille_ko=fichier.size // 1024,
        )
        return Response({'message': 'Pièce jointe ajoutée.', 'id': pj.id})

    @action(detail=True, methods=['get'], url_path='imprimer',
            permission_classes=[IsAgentInstitutionnel])
    def imprimer(self, request, pk=None):
        """GET /api/demandes/{id}/imprimer/ — marque le dossier comme disponible pour impression."""
        demande = self.get_object()
        statut_avant = demande.statut
        # Marquer le statut
        if demande.statut == 'EN_RECEPTION':
            demande.statut = 'PRET_IMPRESSION_DGI'
            demande.save()
        # Enregistrer l'étape
        EtapeTraitement.objects.create(
            demande=demande,
            acteur=request.user,
            etape_code='DGI_SEC_IMPRESSION',
            statut_avant=statut_avant,
            statut_apres=demande.statut,
            action='Dossier mis à disposition pour impression — circuit papier',
            commentaire=request.data.get('commentaire', ''),
        )
        # Retourner les données du dossier pour impression
        from api.serializers import DemandeDetailSerializer
        return Response({
            'message': 'Dossier marqué pour impression.',
            'demande': DemandeDetailSerializer(demande, context={'request': request}).data,
        })

        fichier = request.FILES.get('fichier')
        if not fichier:
            return Response({'error': 'Aucun fichier fourni.'}, status=400)

        piece = PieceJointe.objects.create(
            demande=demande,
            piece_requise_id=request.data.get('piece_requise_id'),
            nom_original=fichier.name,
            fichier=fichier,
            taille_ko=fichier.size // 1024,
        )
        return Response(PieceJointeSerializer(piece).data, status=201)


class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    """Notifications de l'utilisateur connecté."""
    serializer_class   = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Notification.objects.filter(destinataire=self.request.user)
        return qs.order_by('-created_at')

    @action(detail=True, methods=['post'], url_path='marquer-lu')
    def marquer_lu(self, request, pk=None):
        notif = self.get_object()
        notif.lu = True; notif.save()
        return Response({'lu': True})

    @action(detail=False, methods=['post'], url_path='tout-lire')
    def tout_lire(self, request):
        self.get_queryset().update(lu=True)
        return Response({'message': 'Toutes les notifications marquées comme lues.'})

    @action(detail=False, methods=['get'], url_path='non-lues-count')
    def non_lues_count(self, request):
        """GET /api/notifications/non-lues-count/ — nombre de notifications non lues."""
        count = self.get_queryset().filter(lu=False).count()
        return Response({'count': count})


# ─────────────────────────────────────────────────────────────
# AUTORISATIONS
# ─────────────────────────────────────────────────────────────

class AutorisationViewSet(viewsets.ModelViewSet):
    queryset           = Autorisation.objects.all()
    serializer_class   = AutorisationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend]
    filterset_fields   = ['type', 'statut', 'wilaya']

    @action(detail=False, methods=['get'], url_path='geojson', permission_classes=[AllowAny])
    def geojson(self, request):
        """
        GET /api/autorisations/geojson/
        Données publiques pour la cartographie Leaflet.
        """
        qs = Autorisation.objects.filter(
            statut='active', latitude__isnull=False, longitude__isnull=False
        )
        features = []
        for a in qs:
            features.append({
                'type': 'Feature',
                'geometry': {'type': 'Point', 'coordinates': [a.longitude, a.latitude]},
                'properties': {
                    'id': a.id, 'numero_auto': a.numero_auto,
                    'type': a.type, 'wilaya': a.wilaya,
                    'adresse': a.adresse, 'date_delivrance': str(a.date_delivrance),
                }
            })
        return Response({'type': 'FeatureCollection', 'features': features})

    @action(detail=False, methods=['get'], url_path='stats', permission_classes=[AllowAny])
    def stats(self, request):
        """GET /api/autorisations/stats/ — statistiques pour le panel de la carte."""
        from django.db.models import Count
        qs = Autorisation.objects.filter(statut='active')
        total = qs.count()
        par_type = qs.values('type').annotate(count=Count('id'))
        return Response({'total': total, 'par_type': list(par_type)})


class DocumentGenereViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class   = DocumentGenereSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.has_role('DEMANDEUR'):
            return DocumentGenere.objects.filter(
                demande__demandeur=user, signe=True
            )
        return DocumentGenere.objects.all()

    @action(detail=True, methods=['post'], url_path='signer',
            permission_classes=[IsDGIDirecteur])
    def signer(self, request, pk=None):
        """POST /api/documents/{id}/signer/ — signature numérique du document."""
        doc = self.get_object()
        doc.signe = True
        doc.signataire = request.user
        doc.signataire_role = request.user.get_roles()[0] if request.user.get_roles() else ''
        doc.date_signature = timezone.now()
        doc.save()
        return Response({'message': 'Document signé.', 'date_signature': doc.date_signature})


# ─────────────────────────────────────────────────────────────
# MÉTIER SPÉCIFIQUE
# ─────────────────────────────────────────────────────────────

class VisiteLieuxViewSet(viewsets.ModelViewSet):
    queryset           = VisiteLieux.objects.all()
    serializer_class   = VisiteLieuxSerializer
    permission_classes = [IsDDPI]


class ComiteBPViewSet(viewsets.ModelViewSet):
    queryset           = ComiteBP.objects.all()
    serializer_class   = ComiteBPSerializer
    permission_classes = [IsDDPI]


class DistanceBoulangerieViewSet(viewsets.ModelViewSet):
    queryset           = DistanceBoulangerie.objects.all()
    serializer_class   = DistanceBoulangerieSerializer
    permission_classes = [IsDDPI]


class RenouvellementDetailViewSet(viewsets.ModelViewSet):
    queryset           = RenouvellementDetail.objects.all()
    serializer_class   = RenouvellementDetailSerializer
    permission_classes = [IsAuthenticated]


class ExtensionDetailViewSet(viewsets.ModelViewSet):
    queryset           = ExtensionDetail.objects.all()
    serializer_class   = ExtensionDetailSerializer
    permission_classes = [IsAuthenticated]


class UsineEauDetailViewSet(viewsets.ModelViewSet):
    queryset           = UsineEauDetail.objects.all()
    serializer_class   = UsineEauDetailSerializer
    permission_classes = [IsAuthenticated]


class UniteIndustrielleDetailViewSet(viewsets.ModelViewSet):
    queryset           = UniteIndustrielleDetail.objects.all()
    serializer_class   = UniteIndustrielleDetailSerializer
    permission_classes = [IsAuthenticated]


# ─────────────────────────────────────────────────────────────
# ANALYTICS (tableau de bord)
# ─────────────────────────────────────────────────────────────

class AnalyticsView(generics.GenericAPIView):
    """GET /api/analytics/dashboard/ — données pour le tableau de bord."""
    permission_classes = [IsAgentInstitutionnel]

    def get(self, request):
        from django.db.models import Count
        user = request.user

        qs_demandes = Demande.objects.all()
        if user.has_role('DEMANDEUR'):
            qs_demandes = qs_demandes.filter(demandeur=user)

        total         = qs_demandes.count()
        par_statut    = list(qs_demandes.values('statut').annotate(count=Count('id')))
        par_type      = list(qs_demandes.values('type_demande__code').annotate(count=Count('id')))
        par_wilaya    = list(qs_demandes.values('wilaya').annotate(count=Count('id')).order_by('-count')[:10])
        recentes      = DemandeListSerializer(
            qs_demandes.order_by('-date_soumission')[:5], many=True
        ).data

        return Response({
            'total_demandes':  total,
            'par_statut':      par_statut,
            'par_type':        par_type,
            'par_wilaya':      par_wilaya,
            'demandes_recentes': recentes,
            'autorisations_actives': Autorisation.objects.filter(statut='active').count(),
        })


class LoginAgentView(generics.GenericAPIView):
    """
    POST /api/auth/login/agent/
    Connexion des agents et administrateurs via email + password.
    Les demandeurs utilisent /api/auth/login/ avec identifiant_unique.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        email    = request.data.get('email', '').strip().lower()
        password = request.data.get('password', '')

        if not email or not password:
            return Response({'detail': 'Email et mot de passe obligatoires.'}, status=400)

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({'detail': 'Email ou mot de passe incorrect.'}, status=401)

        if not user.check_password(password):
            return Response({'detail': 'Email ou mot de passe incorrect.'}, status=401)

        if not user.is_active:
            return Response({'detail': 'Compte désactivé. Contactez l\'administrateur.'}, status=403)

        # Bloquer les demandeurs — ils ont leur propre page de connexion
        roles = user.get_roles()
        if roles == ['DEMANDEUR']:
            return Response(
                {'detail': 'Les demandeurs utilisent l\'espace demandeur avec leur identifiant MMI-DEM-XXXXX.'},
                status=403
            )

        from rest_framework_simplejwt.tokens import RefreshToken
        refresh = RefreshToken.for_user(user)

        # Mettre à jour last_login
        from django.utils import timezone
        user.last_login_at = timezone.now()
        user.save(update_fields=['last_login_at'])

        from .serializers import UserDetailSerializer
        return Response({
            'access':  str(refresh.access_token),
            'refresh': str(refresh),
            'user':    UserDetailSerializer(user).data,
        })


class PasswordChangeView(generics.GenericAPIView):
    """POST /api/auth/password-change/ — changer son propre mot de passe."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        old  = request.data.get('old_password', '')
        new  = request.data.get('new_password', '')
        if not user.check_password(old):
            return Response({'detail': 'Mot de passe actuel incorrect.'}, status=400)
        if len(new) < 8:
            return Response({'detail': 'Le mot de passe doit contenir au moins 8 caractères.'}, status=400)
        user.set_password(new)
        user.save()
        return Response({'detail': 'Mot de passe modifié avec succès.'})


# ─────────────────────────────────────────────────────────────
# RESET MOT DE PASSE (lien email)
# ─────────────────────────────────────────────────────────────

class ResetPasswordConfirmView(generics.GenericAPIView):
    """POST /api/auth/reset-password-confirm/ — confirme le nouveau mot de passe via token."""
    permission_classes = [AllowAny]

    def post(self, request):
        uidb64   = request.data.get('uidb64', '')
        token    = request.data.get('token', '')
        password = request.data.get('password', '')

        if not all([uidb64, token, password]):
            return Response({'detail': 'Données manquantes.'}, status=400)

        if len(password) < 8:
            return Response({'detail': 'Le mot de passe doit contenir au moins 8 caractères.'}, status=400)

        try:
            uid  = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            return Response({'detail': 'Lien invalide ou expiré.'}, status=400)

        if not default_token_generator.check_token(user, token):
            return Response({'detail': 'Lien expiré ou déjà utilisé.'}, status=400)

        user.set_password(password)
        user.save()
        return Response({'detail': 'Mot de passe réinitialisé avec succès.'})


class PasswordResetRequestView(generics.GenericAPIView):
    """POST /api/auth/password-reset/ — demande de réinitialisation par email."""
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email', '').strip().lower()
        try:
            user = User.objects.get(email=email, is_active=True)
            from django.core.mail import send_mail
            uid   = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            reset_url = f"{settings.FRONTEND_URL}/reset-password/{uid}/{token}/"
            send_mail(
                subject="Réinitialisation de votre mot de passe — MMI",
                message=(
                    f"Bonjour {user.nom_complet},\n\n"
                    f"Cliquez sur ce lien pour réinitialiser votre mot de passe :\n\n"
                    f"{reset_url}\n\n"
                    f"Ce lien est valable 72 heures.\n\n"
                    f"Si vous n\'avez pas demandé cette réinitialisation, ignorez cet email.\n\n"
                    f"Cordialement,\nL\'équipe MMI"
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )
        except User.DoesNotExist:
            pass
        return Response({'detail': 'Si cet email existe, un lien de réinitialisation a été envoyé.'})


# ─────────────────────────────────────────────────────────────
# ANALYTICS AVANCÉ — DGI + Export
# ─────────────────────────────────────────────────────────────

class DGIAnalyticsView(generics.GenericAPIView):
    """GET /api/analytics/dgi/ — tableau de bord complet DGI avec données renouvellement."""
    permission_classes = [IsAgentInstitutionnel]

    def get(self, request):
        from django.db.models import Count, Q, Avg
        from django.utils import timezone
        from datetime import timedelta

        today = timezone.now().date()
        debut_annee = today.replace(month=1, day=1)

        qs = Demande.objects.all()

        # KPIs généraux
        kpis = {
            'total_demandes':        qs.count(),
            'en_cours':              qs.exclude(statut__in=['VALIDE','REJETE']).count(),
            'valides_annee':         qs.filter(statut='VALIDE', date_soumission__date__gte=debut_annee).count(),
            'rejetes':               qs.filter(statut='REJETE').count(),
            'autorisations_actives': Autorisation.objects.filter(statut='active').count(),
            'en_instruction_dgi':    qs.filter(statut='EN_INSTRUCTION_DGI').count(),
            'en_attente_signature':  qs.filter(statut='SIGNATURE_DGI').count(),
        }

        # Par type de demande
        par_type = list(
            qs.values('type_demande__code', 'type_demande__libelle')
              .annotate(total=Count('id'),
                        valides=Count('id', filter=Q(statut='VALIDE')),
                        en_cours=Count('id', filter=~Q(statut__in=['VALIDE','REJETE'])))
              .order_by('-total')
        )

        # Par wilaya
        par_wilaya = list(
            qs.exclude(wilaya='').values('wilaya')
              .annotate(total=Count('id'))
              .order_by('-total')[:15]
        )

        # Par mois (12 derniers mois)
        par_mois = []
        for i in range(11, -1, -1):
            d = today - timedelta(days=30*i)
            count = qs.filter(
                date_soumission__year=d.year,
                date_soumission__month=d.month
            ).count()
            par_mois.append({
                'mois': d.strftime('%b %Y'),
                'annee': d.year,
                'mois_num': d.month,
                'count': count,
            })

        # Par statut
        par_statut = list(qs.values('statut').annotate(count=Count('id')).order_by('-count'))

        # Dernières demandes
        recentes = DemandeListSerializer(
            qs.order_by('-date_soumission')[:10], many=True
        ).data

        # ── Données renouvellement ────────────────────────────
        from api.models import RenouvellementDetail
        renouvellements = RenouvellementDetail.objects.select_related('demande__demandeur','demande__type_demande')

        renouv_data = []
        for r in renouvellements[:50]:
            renouv_data.append({
                'id':                    r.id,
                'numero_ref':            r.demande.numero_ref,
                'raison_sociale':        r.demande.raison_sociale,
                'statut_demande':        r.demande.statut,
                'wilaya':                r.demande.wilaya,
                'activite':              r.demande.activite,
                'date_soumission':       r.demande.date_soumission.strftime('%d/%m/%Y'),
                # Section I
                'abreviation':           r.abreviation,
                'nationalite_entreprise':r.nationalite_entreprise,
                # Section III
                'nom_responsable':       r.nom_responsable,
                'nni_passeport':         r.nni_passeport,
                'telephone_responsable': r.telephone_responsable,
                # Section IV
                'forme_juridique':       r.forme_juridique,
                'registre_commerce':     r.registre_commerce,
                'nif_numero':            r.nif_numero,
                'cnss_numero':           r.cnss_numero,
                # Section V
                'numero_enregistrement': r.numero_enregistrement,
                'date_creation':         str(r.date_creation) if r.date_creation else '',
                'date_debut_production': str(r.date_debut_production) if r.date_debut_production else '',
                'emplois_crees':         r.emplois_crees,
                'employes_administratifs':r.employes_administratifs,
                'techniciens':           r.techniciens,
                'ouvriers':              r.ouvriers,
                # Section VI
                'description_unite':     r.description_unite,
                'capital_social':        str(r.capital_social) if r.capital_social else '',
                'capacite_tonnes_an':    r.capacite_tonnes_an,
                'capacite_tonnes_jour':  r.capacite_tonnes_jour,
                'varietes_production':   r.varietes_production,
                'difficultes':           r.difficultes,
            })

        # Stats renouvellement
        renouv_stats = {
            'total':       renouvellements.count(),
            'par_forme':   list(renouvellements.values('forme_juridique').annotate(count=Count('id'))),
            'total_emplois': sum(r.emplois_crees for r in renouvellements),
        }

        return Response({
            'kpis':            kpis,
            'par_type':        par_type,
            'par_wilaya':      par_wilaya,
            'par_mois':        par_mois,
            'par_statut':      par_statut,
            'demandes_recentes': recentes,
            'renouvellements': renouv_data,
            'renouv_stats':    renouv_stats,
        })


class ExportView(generics.GenericAPIView):
    """
    GET /api/export/renouvellements/?format=csv|excel
    Export CSV ou Excel des données de renouvellement pour le DGI.
    """
    permission_classes = [IsAuthenticated]  # Ouvert à tous les agents authentifiés

    def get(self, request):
        import csv
        import io
        from django.http import HttpResponse
        from api.models import RenouvellementDetail

        fmt = request.query_params.get('format', 'csv')
        qs  = RenouvellementDetail.objects.select_related(
            'demande', 'demande__demandeur', 'demande__type_demande'
        ).order_by('-demande__date_soumission')

        # Filtres optionnels
        wilaya = request.query_params.get('wilaya')
        statut = request.query_params.get('statut')
        annee  = request.query_params.get('annee')
        if wilaya: qs = qs.filter(demande__wilaya=wilaya)
        if statut: qs = qs.filter(demande__statut=statut)
        if annee:  qs = qs.filter(demande__date_soumission__year=annee)

        HEADERS = [
            'N° Référence','Raison Sociale','Activité','Wilaya','Statut','Date Soumission',
            'Abréviation','Nationalité Entreprise',
            'Nom Responsable','NNI/Passeport','Tél Responsable',
            'Forme Juridique','N° RC','NIF','CNSS',
            'N° Enregistrement','Date Création','Date Début Production','Date Démarrage',
            'Emplois Créés','Employés Admin','Techniciens','Ouvriers','Total Employés',
            'Nationalités Employés',
            'Description Unité','Capital Social (MRU)','Fonds Propres','Emprunt',
            'Nature Investissement','Type Données',
            'Capacité (T/jour)','Capacité (T/mois)','Capacité (T/an)',
            'Production Effective','Stock',
            'Variétés Production','Capacité Estimée','Capacité Augmentation',
            'Difficultés',
        ]

        def row(r):
            total_emp = (r.employes_administratifs or 0) + (r.techniciens or 0) + (r.ouvriers or 0)
            return [
                r.demande.numero_ref,
                r.demande.raison_sociale,
                r.demande.activite,
                r.demande.wilaya,
                r.demande.get_statut_display(),
                r.demande.date_soumission.strftime('%d/%m/%Y'),
                r.abreviation,
                r.nationalite_entreprise,
                r.nom_responsable,
                r.nni_passeport,
                r.telephone_responsable,
                r.get_forme_juridique_display() if r.forme_juridique else '',
                r.registre_commerce,
                r.nif_numero,
                r.cnss_numero,
                r.numero_enregistrement,
                str(r.date_creation) if r.date_creation else '',
                str(r.date_debut_production) if r.date_debut_production else '',
                str(r.date_demarrage) if r.date_demarrage else '',
                r.emplois_crees,
                r.employes_administratifs,
                r.techniciens,
                r.ouvriers,
                total_emp,
                r.nationalites_employes,
                r.description_unite,
                str(r.capital_social or ''),
                str(r.fonds_propres or ''),
                str(r.emprunt or ''),
                r.nature_investissement,
                r.type_donnees,
                r.capacite_tonnes_jour,
                r.capacite_tonnes_mois,
                r.capacite_tonnes_an,
                r.production_effective,
                r.stock,
                r.varietes_production,
                r.capacite_augmentation,
                ' | '.join(r.difficultes) if isinstance(r.difficultes, list) else str(r.difficultes),
            ]

        if fmt == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="renouvellements_mmi.csv"'
            writer = csv.writer(response)
            writer.writerow(HEADERS)
            for r in qs:
                writer.writerow(row(r))
            return response

        elif fmt == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Renouvellements MMI"

                # En-tête entreprise
                ws.merge_cells('A1:AK1')
                ws['A1'] = "MINISTÈRE DES MINES ET DE L'INDUSTRIE — Export Renouvellements Industriels"
                ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
                ws['A1'].fill = PatternFill('solid', fgColor='1B6B30')
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 30

                # En-têtes colonnes
                header_fill = PatternFill('solid', fgColor='2E8B45')
                header_font = Font(bold=True, color='FFFFFF', size=10)
                for col_idx, h in enumerate(HEADERS, 1):
                    cell = ws.cell(row=2, column=col_idx, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(h) + 2)
                ws.row_dimensions[2].height = 40

                # Données
                alt_fill = PatternFill('solid', fgColor='F0FFF4')
                for row_idx, r in enumerate(qs, 3):
                    for col_idx, val in enumerate(row(r), 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        if row_idx % 2 == 0:
                            cell.fill = alt_fill

                # Figer la première ligne
                ws.freeze_panes = 'A3'

                # Filtre automatique
                ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}2"

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                response = HttpResponse(
                    buf.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="renouvellements_mmi.xlsx"'
                return response
            except ImportError:
                # openpyxl manquant — installer automatiquement
                import subprocess, sys
                try:
                    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'], 
                                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    return Response({'detail': 'openpyxl installé. Relancez l\'export.'}, status=503)
                except Exception:
                    return Response({'detail': 'openpyxl non disponible. Utilisez format=csv'}, status=501)

        return Response({'detail': 'Format invalide. Utilisez format=csv ou format=excel'}, status=400)


# ─────────────────────────────────────────────────────────────
# SUPER ADMIN — Gestion complète
# ─────────────────────────────────────────────────────────────

class AdminActualiteView(generics.GenericAPIView):
    """CRUD Actualités — /api/admin/actualites/"""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        from api.models import Actualite
        from api.serializers import ActualiteSerializer
        qs = Actualite.objects.all().order_by('-date_publication')
        search = request.query_params.get('search')
        if search:
            qs = qs.filter(titre__icontains=search)
        return Response({'results': ActualiteSerializer(qs, many=True).data, 'count': qs.count()})

    def post(self, request):
        from api.models import Actualite
        data = request.data
        # publie peut arriver comme string "true"/"false" via multipart
        publie_val = data.get('publie', False)
        if isinstance(publie_val, str):
            publie_val = publie_val.lower() in ('true', '1', 'yes')
        art = Actualite(
            titre=data.get('titre', ''),
            contenu=data.get('contenu', ''),
            publie=publie_val,
            auteur=request.user,
        )
        # Gérer l'image si envoyée en multipart
        if 'image' in request.FILES:
            art.image = request.FILES['image']
        art.save()
        return Response({'id': art.id, 'titre': art.titre, 'slug': art.slug}, status=201)


class AdminActualiteDetailView(generics.GenericAPIView):
    """PATCH/DELETE /api/admin/actualites/<id>/"""
    permission_classes = [IsSuperAdmin]

    def patch(self, request, pk):
        from api.models import Actualite
        try:
            art = Actualite.objects.get(pk=pk)
        except Actualite.DoesNotExist:
            return Response({'detail': 'Introuvable'}, status=404)
        for field in ['titre', 'contenu']:
            if field in request.data:
                setattr(art, field, request.data[field])
        # publie peut arriver comme string via multipart
        if 'publie' in request.data:
            val = request.data['publie']
            if isinstance(val, str):
                val = val.lower() in ('true', '1', 'yes')
            art.publie = val
        # Image
        if 'image' in request.FILES:
            art.image = request.FILES['image']
        art.save()
        return Response({'detail': 'Modifié'})

    def delete(self, request, pk):
        from api.models import Actualite
        try:
            Actualite.objects.get(pk=pk).delete()
            return Response({'detail': 'Supprimé'})
        except Actualite.DoesNotExist:
            return Response({'detail': 'Introuvable'}, status=404)


class AdminDocumentView(generics.GenericAPIView):
    """CRUD Documents publics — /api/admin/documents/"""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        from api.models import DocumentPublic
        qs = DocumentPublic.objects.all()
        cat = request.query_params.get('categorie')
        if cat:
            qs = qs.filter(categorie=cat)
        data = [{'id': d.id, 'titre': d.titre, 'categorie': d.categorie,
                 'publie': d.publie, 'fichier': d.fichier.url if d.fichier else None,
                 'created_at': d.created_at.strftime('%d/%m/%Y')} for d in qs]
        return Response({'results': data, 'count': len(data)})

    def post(self, request):
        from api.models import DocumentPublic
        doc = DocumentPublic.objects.create(
            titre=request.data.get('titre', ''),
            categorie=request.data.get('categorie', 'ANNEXE'),
            fichier=request.FILES.get('fichier'),
            publie=request.data.get('publie', False),
            created_by=request.user,
        )
        return Response({'id': doc.id, 'titre': doc.titre}, status=201)

    def delete(self, request, pk=None):
        from api.models import DocumentPublic
        try:
            DocumentPublic.objects.get(pk=pk).delete()
            return Response({'detail': 'Supprimé'})
        except DocumentPublic.DoesNotExist:
            return Response({'detail': 'Introuvable'}, status=404)


class AdminStatsView(generics.GenericAPIView):
    """GET /api/admin/stats/ — statistiques globales superadmin"""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        from django.db.models import Count, Q
        from django.utils import timezone
        from datetime import timedelta

        today = timezone.now()
        debut_semaine = today - timedelta(days=7)
        debut_mois    = today - timedelta(days=30)

        users_qs   = User.objects.all()
        demandes_qs = Demande.objects.all()

        return Response({
            # Utilisateurs
            'total_users':       users_qs.count(),
            'demandeurs':        users_qs.filter(user_roles__role__code='DEMANDEUR', user_roles__actif=True).count(),
            'agents':            users_qs.filter(is_staff=False, is_super_admin=False).exclude(user_roles__role__code='DEMANDEUR').count(),
            'users_actifs':      users_qs.filter(is_active=True).count(),
            'users_suspendus':   users_qs.filter(is_active=False).count(),
            'nouveaux_7j':       users_qs.filter(created_at__gte=debut_semaine).count(),
            # Demandes
            'total_demandes':    demandes_qs.count(),
            'demandes_semaine':  demandes_qs.filter(date_soumission__gte=debut_semaine).count(),
            'demandes_mois':     demandes_qs.filter(date_soumission__gte=debut_mois).count(),
            'en_cours':          demandes_qs.exclude(statut__in=['VALIDE', 'REJETE']).count(),
            'valides':           demandes_qs.filter(statut='VALIDE').count(),
            'rejetes':           demandes_qs.filter(statut='REJETE').count(),
            # Par type
            'par_type': list(demandes_qs.values('type_demande__code').annotate(count=Count('id'))),
            # Par statut
            'par_statut': list(demandes_qs.values('statut').annotate(count=Count('id')).order_by('-count')),
            # Autorisations
            'autorisations_actives': Autorisation.objects.filter(statut='active').count(),
            # Roles
            'par_role': list(
                UserRole.objects.filter(actif=True)
                .values('role__code', 'role__nom')
                .annotate(count=Count('id'))
                .order_by('-count')
            ),
        })


class AdminUserDetailView(generics.GenericAPIView):
    """PATCH /api/admin/users/<id>/ — modifier un utilisateur"""
    permission_classes = [IsSuperAdmin]

    def patch(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({'detail': 'Introuvable'}, status=404)

        for field in ['nom', 'prenom', 'email', 'telephone', 'is_active']:
            if field in request.data:
                setattr(user, field, request.data[field])

        if 'password' in request.data and request.data['password']:
            user.set_password(request.data['password'])

        user.save()

        # Changer le rôle si demandé
        if 'role_code' in request.data:
            from api.models import Role, UserRole
            try:
                role = Role.objects.get(code=request.data['role_code'])
                ancien_roles = list(UserRole.objects.filter(user=user).values_list('role__code', flat=True))
                # Supprimer TOUS les anciens rôles (même logique que attribuer_role)
                UserRole.objects.filter(user=user).delete()
                # Créer le nouveau rôle unique et actif
                UserRole.objects.create(
                    user=user, role=role,
                    actif=True, assigned_by=request.user
                )
                logger.info(
                    f"Rôle de {user.email} changé : {ancien_roles} → {role.code} "
                    f"par {request.user.email}"
                )
            except Role.DoesNotExist:
                return Response({'detail': f'Rôle « {request.data["role_code"]} » introuvable.'}, status=400)

        return Response({'detail': 'Utilisateur modifié avec succès'})

    def delete(self, request, pk):
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({'detail': 'Introuvable'}, status=404)

        # Protections
        if user == request.user:
            return Response({'detail': 'Vous ne pouvez pas supprimer votre propre compte.'}, status=403)
        if user.is_super_admin:
            return Response({'detail': 'Impossible de supprimer un super administrateur.'}, status=403)

        # Mode de suppression selon le paramètre
        mode = request.query_params.get('mode', 'suspend')  # suspend | delete

        if mode == 'delete':
            # Suppression définitive
            nom = user.nom_complet
            user.delete()
            return Response({'detail': f'Utilisateur « {nom} » supprimé définitivement.'})
        else:
            # Suspension (désactivation — conserve les données)
            user.is_active = not user.is_active
            user.save()
            action = 'réactivé' if user.is_active else 'suspendu'
            return Response({'detail': f'Utilisateur {action} avec succès.', 'is_active': user.is_active})


# ─────────────────────────────────────────────────────────────
# ACTIVATION COMPTE AGENT & CHANGEMENT MOT DE PASSE
# ─────────────────────────────────────────────────────────────

class ActivationCompteView(generics.GenericAPIView):
    """GET /api/auth/activer/<uidb64>/<token>/ — active le compte agent."""
    permission_classes = [AllowAny]

    def get(self, request, uidb64, token):
        try:
            uid  = force_str(urlsafe_base64_decode(uidb64))
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            return Response({'detail': 'Lien invalide.'}, status=400)

        if default_token_generator.check_token(user, token):
            if user.is_active:
                return Response({'detail': 'Compte déjà activé.'}, status=200)
            user.is_active = True
            user.save()
            return Response({'detail': 'Compte activé avec succès. Vous pouvez maintenant vous connecter.'})
        return Response({'detail': 'Lien expiré ou invalide.'}, status=400)


class PasswordChangeView(generics.GenericAPIView):
    """POST /api/auth/password-change/ — change le mot de passe (connecté)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ancien   = request.data.get('ancien_password', '')
        nouveau  = request.data.get('nouveau_password', '')
        if not request.user.check_password(ancien):
            return Response({'detail': 'Ancien mot de passe incorrect.'}, status=400)
        if len(nouveau) < 8:
            return Response({'detail': 'Le nouveau mot de passe doit contenir au moins 8 caractères.'}, status=400)
        request.user.set_password(nouveau)
        request.user.save()
        return Response({'detail': 'Mot de passe modifié avec succès.'})



class QuittanceBPView(generics.GenericAPIView):
    """
    POST /api/demandes/{id}/quittance-bp/
    Secrétaire comité BP dépose la quittance de paiement après approbation PV.
    """
    permission_classes = [IsAgentInstitutionnel]

    def post(self, request, pk=None):
        from api.models import Demande, EtapeTraitement
        try:
            demande = Demande.objects.get(pk=pk)
        except Demande.DoesNotExist:
            return Response({'detail': 'Dossier introuvable.'}, status=404)

        if demande.statut not in ['PV_COMITE_DEPOSE', 'ATTENTE_QUITTANCE']:
            return Response({'detail': 'Ce dossier n\'est pas en attente de quittance.'}, status=400)

        if 'quittance' not in request.FILES:
            return Response({'detail': 'Veuillez joindre le fichier de quittance.'}, status=400)

        statut_avant = demande.statut
        demande.statut = 'ATTENTE_QUITTANCE'
        demande.save()

        etape = EtapeTraitement.objects.create(
            demande=demande,
            acteur=request.user,
            etape_code='DDPI_QUITTANCE_BP',
            statut_avant=statut_avant,
            statut_apres='ATTENTE_QUITTANCE',
            action='Quittance de paiement reçue et enregistrée',
            commentaire=request.data.get('commentaire', ''),
        )
        etape.piece_jointe = request.FILES['quittance']
        etape.save()

        _notifier_prochaine_etape(demande, 'DDPI_QUITTANCE_BP', request.user)
        return Response({'message': 'Quittance enregistrée.', 'etape_id': etape.id})


class PVComiteBPView(generics.GenericAPIView):
    """
    POST /api/demandes/{id}/pv-comite-bp/
    Secrétaire comité BP dépose le PV d\'approbation du comité.
    """
    permission_classes = [IsAgentInstitutionnel]

    def post(self, request, pk=None):
        from api.models import Demande, EtapeTraitement
        try:
            demande = Demande.objects.get(pk=pk)
        except Demande.DoesNotExist:
            return Response({'detail': 'Dossier introuvable.'}, status=404)

        if demande.statut != 'EN_COMMISSION_BP':
            return Response({'detail': 'Ce dossier n\'est pas en commission BP.'}, status=400)

        if 'pv' not in request.FILES:
            return Response({'detail': 'Veuillez joindre le fichier PV.'}, status=400)

        approuve = request.data.get('approuve', 'true').lower() == 'true'
        statut_avant = demande.statut

        if approuve:
            demande.statut = 'PV_COMITE_DEPOSE'
        else:
            demande.statut = 'REJETE'

        demande.save()

        etape = EtapeTraitement.objects.create(
            demande=demande,
            acteur=request.user,
            etape_code='DDPI_PV_COMITE_BP',
            statut_avant=statut_avant,
            statut_apres=demande.statut,
            action=f"PV comité BP {'approuvé' if approuve else 'rejeté'} — document déposé",
            commentaire=request.data.get('commentaire', ''),
        )
        etape.piece_jointe = request.FILES['pv']
        etape.save()

        _notifier_prochaine_etape(demande, 'DDPI_PV_COMITE_BP', request.user)
        return Response({
            'message': f"PV {'approuvé' if approuve else 'rejeté'}. Dossier mis à jour.",
            'statut':  demande.statut,
        })


# ─────────────────────────────────────────────────────────────
# CONFIGURATION PLATEFORME
# ─────────────────────────────────────────────────────────────

class ConfigPlateformeView(generics.GenericAPIView):
    """GET/POST /api/admin/config/ — lire et modifier les parametres."""

    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAuthenticated()]
        return [IsSuperAdmin()]

    def get(self, request):
        from api.models import ConfigPlateforme
        return Response({
            'ministre_nom':   ConfigPlateforme.get('ministre_nom',   'Le Ministre'),
            'ministre_titre': ConfigPlateforme.get('ministre_titre',  'Ministre des Mines et de l\'Industrie'),
            'plateforme_url': ConfigPlateforme.get('plateforme_url',  'https://plateforme.mmi.gov.mr'),
        })

    def post(self, request):
        from api.models import ConfigPlateforme
        updated = []
        for cle, valeur in request.data.items():
            obj, _ = ConfigPlateforme.objects.get_or_create(cle=cle)
            obj.valeur = valeur
            obj.updated_by = request.user
            obj.save()
            updated.append(cle)
        return Response({'message': f'{len(updated)} parametre(s) mis a jour.', 'updated': updated})


# ─────────────────────────────────────────────────────────────
# QUITTANCE PAIEMENT (après comité BP)
# ─────────────────────────────────────────────────────────────

class QuittancePaiementView(generics.GenericAPIView):
    """POST /api/demandes/{id}/quittance/ — dépôt quittance paiement BP."""
    permission_classes = [IsAgentInstitutionnel]

    def post(self, request, pk=None):
        from api.models import Demande, QuittancePaiement, EtapeTraitement, Notification
        try:
            demande = Demande.objects.get(pk=pk)
        except Demande.DoesNotExist:
            return Response({'detail': 'Demande introuvable.'}, status=404)

        if demande.statut not in ['PV_COMITE_DEPOSE', 'EN_COMMISSION_BP', 'ATTENTE_QUITTANCE']:
            return Response({'detail': 'La quittance ne peut etre deposee qu apres approbation du comite BP.'}, status=400)

        fichier = request.FILES.get('fichier')
        if not fichier:
            return Response({'detail': 'Fichier quittance obligatoire.'}, status=400)

        q, created = QuittancePaiement.objects.get_or_create(
            demande=demande,
            defaults={
                'fichier':    fichier,
                'montant':    request.data.get('montant'),
                'reference':  request.data.get('reference', ''),
                'depose_par': request.user,
                'valide':     True,
            }
        )
        if not created:
            q.fichier   = fichier
            q.montant   = request.data.get('montant')
            q.reference = request.data.get('reference', '')
            q.valide    = True
            q.save()

        statut_avant = demande.statut
        demande.statut = 'ACCORD_PRINCIPE'
        demande.save()

        EtapeTraitement.objects.create(
            demande=demande, acteur=request.user,
            etape_code='DDPI_QUITTANCE_BP',
            statut_avant=statut_avant, statut_apres='ACCORD_PRINCIPE',
            action='Quittance de paiement deposee — Accord de principe accorde',
            commentaire=request.data.get('commentaire', ''),
        )

        Notification.objects.create(
            destinataire=demande.demandeur, demande=demande,
            type_notif='ACCORD_PRINCIPE',
            titre='Accord de principe accordé',
            message=f'Votre demande {demande.numero_ref} a recu un accord de principe apres validation du comite BP et reception de votre quittance de paiement.',
        )
        return Response({'message': 'Quittance deposee — Accord de principe accorde.', 'statut': 'ACCORD_PRINCIPE'})


class PVComiteBPView(generics.GenericAPIView):
    """POST /api/demandes/{id}/pv-comite/ — depot PV comite BP."""
    permission_classes = [IsAgentInstitutionnel]

    def post(self, request, pk=None):
        from api.models import Demande, EtapeTraitement, Notification
        try:
            demande = Demande.objects.get(pk=pk)
        except Demande.DoesNotExist:
            return Response({'detail': 'Demande introuvable.'}, status=404)

        statut_avant = demande.statut
        demande.statut = 'PV_COMITE_DEPOSE'
        demande.save()

        etape = EtapeTraitement(
            demande=demande, acteur=request.user,
            etape_code='DDPI_PV_COMITE_BP',
            statut_avant=statut_avant, statut_apres='PV_COMITE_DEPOSE',
            action='PV du comite BP depose — En attente quittance paiement',
            commentaire=request.data.get('commentaire', ''),
        )
        if 'piece_jointe' in request.FILES:
            etape.piece_jointe = request.FILES['piece_jointe']
        etape.save()

        Notification.objects.create(
            destinataire=demande.demandeur, demande=demande,
            type_notif='ACCORD_PRINCIPE',
            titre='Comite BP - Resultat favorable',
            message=f'Votre demande {demande.numero_ref} a ete approuvee par le comite BP. Veuillez deposer votre quittance de paiement pour finaliser.',
        )
        return Response({'message': 'PV depose. En attente quittance.', 'statut': 'PV_COMITE_DEPOSE'})


class ImprimerDossierView(generics.GenericAPIView):
    """GET /api/demandes/{id}/imprimer/ — genere un recapitulatif imprimable du dossier."""
    permission_classes = [IsAgentInstitutionnel]

    def get(self, request, pk=None):
        from api.models import Demande
        try:
            demande = Demande.objects.get(pk=pk)
        except Demande.DoesNotExist:
            return Response({'detail': 'Introuvable.'}, status=404)

        from api.serializers import DemandeDetailSerializer
        data = DemandeDetailSerializer(demande, context={'request': request}).data
        return Response({
            'demande':   data,
            'print_url': f"/api/demandes/{pk}/imprimer/",
            'message':   'Donnees pour impression disponibles.',
        })


class BackupDatabaseView(generics.GenericAPIView):
    """GET /api/admin/backup/ — télécharge la base de données SQLite3."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        import os, mimetypes
        from django.http import FileResponse
        from django.conf import settings as dj_settings

        db_path = dj_settings.DATABASES['default']['NAME']
        if not os.path.exists(db_path):
            return Response({'detail': 'Base de données introuvable.'}, status=404)

        import datetime
        filename = f"mmi_backup_{datetime.date.today().strftime('%Y%m%d')}.sqlite3"

        response = FileResponse(
            open(db_path, 'rb'),
            content_type='application/octet-stream',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ContactView(generics.GenericAPIView):
    """POST /api/contact/ — formulaire de contact public."""
    permission_classes = [AllowAny]

    def post(self, request):
        nom       = request.data.get('nom', '').strip()
        email     = request.data.get('email', '').strip()
        telephone = request.data.get('telephone', '').strip()
        sujet     = request.data.get('sujet', '').strip()
        message   = request.data.get('message', '').strip()

        if not all([nom, email, message]):
            return Response({'detail': 'Nom, email et message sont requis.'}, status=400)

        from django.core.mail import send_mail
        try:
            send_mail(
                subject=f'[MMI Contact] {sujet} — {nom}',
                message=(
                    f'Nouveau message depuis le formulaire de contact MMI\n\n'
                    f'Nom       : {nom}\n'
                    f'Email     : {email}\n'
                    f'Téléphone : {telephone or "Non renseigné"}\n'
                    f'Sujet     : {sujet}\n\n'
                    f'Message :\n{message}\n'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[settings.DEFAULT_FROM_EMAIL],
                reply_to=[email],
                fail_silently=False,
            )
            # Email de confirmation au visiteur
            send_mail(
                subject='Votre message a été reçu — Ministère des Mines et de l\'Industrie',
                message=(
                    f'Bonjour {nom},\n\n'
                    f'Nous avons bien reçu votre message et vous répondrons dans les meilleurs délais.\n\n'
                    f'Sujet : {sujet}\n\n'
                    f'Cordialement,\nMinistère des Mines et de l\'Industrie — Mauritanie'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )
        except Exception as exc:
            logger.warning(f'Email contact échoué: {exc}')

        return Response({'detail': 'Message envoyé avec succès.'})


class AutorisationByNumeroView(generics.GenericAPIView):
    """GET /api/autorisations/by-numero/?numero=AUTO-XXX&pour=renouvellement|extension — chercher une autorisation."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date as d
        numero = request.query_params.get('numero', '').strip()
        pour   = request.query_params.get('pour', 'extension')  # renouvellement ou extension

        if not numero:
            return Response({'detail': 'Paramètre numero requis.'}, status=400)

        try:
            auto = Autorisation.objects.get(numero_auto=numero)
        except Autorisation.DoesNotExist:
            return Response({'detail': f'Aucune autorisation trouvée avec le numéro « {numero} ».'}, status=404)

        today = d.today()
        est_expiree = auto.date_expiration and auto.date_expiration < today

        # Pour un RENOUVELLEMENT : l'autorisation doit être expirée OU proche de l'expiration (≤ 90 jours)
        if pour == 'renouvellement':
            if auto.statut == 'revoquee':
                return Response({'detail': 'Cette autorisation a été révoquée et ne peut pas être renouvelée.'}, status=400)
            if not auto.date_expiration:
                return Response({'detail': 'Cette autorisation est permanente et ne nécessite pas de renouvellement.'}, status=400)

            jours_restants = (auto.date_expiration - today).days if not est_expiree else 0

            return Response({
                'id':              auto.id,
                'numero_auto':     auto.numero_auto,
                'type':            auto.type,
                'wilaya':          auto.wilaya,
                'adresse':         auto.adresse,
                'date_delivrance': str(auto.date_delivrance),
                'date_expiration': str(auto.date_expiration),
                'statut':          auto.statut,
                'est_expiree':     est_expiree,
                'jours_restants':  jours_restants if not est_expiree else 0,
                'peut_renouveler': est_expiree or jours_restants <= 90,
                'message':         (
                    'Autorisation expirée — renouvellement éligible.' if est_expiree
                    else f'Expire dans {jours_restants} jour(s) — renouvellement possible.'
                    if jours_restants <= 90
                    else f'Autorisation encore valide ({jours_restants} jours). Renouvellement possible à 90 jours de l\'expiration.'
                ),
            })

        # Pour une EXTENSION : l'autorisation doit être active
        else:
            if auto.statut != 'active':
                return Response({'detail': f'Cette autorisation est « {auto.statut} » et ne peut pas être étendue.'}, status=400)
            if est_expiree:
                return Response({'detail': 'Cette autorisation est expirée. Faites d\'abord un renouvellement.'}, status=400)

            return Response({
                'id':              auto.id,
                'numero_auto':     auto.numero_auto,
                'type':            auto.type,
                'wilaya':          auto.wilaya,
                'adresse':         auto.adresse,
                'date_delivrance': str(auto.date_delivrance),
                'date_expiration': str(auto.date_expiration) if auto.date_expiration else None,
                'statut':          auto.statut,
                'peut_etendre':    True,
                'message':         'Autorisation active — extension éligible.',
            })


class LierAutorisationRenouvellementView(generics.GenericAPIView):
    """POST /api/demandes/{id}/lier-autorisation/ — lie une autorisation à un renouvellement."""
    permission_classes = [IsDemandeur]

    def post(self, request, pk):
        from api.models import Autorisation as Auto
        try:
            demande = Demande.objects.get(pk=pk, demandeur=request.user)
        except Demande.DoesNotExist:
            return Response({'detail': 'Demande introuvable.'}, status=404)

        if demande.type_demande.code not in ('RENOUVELLEMENT', 'EXTENSION'):
            return Response({'detail': 'Cette action est réservée aux renouvellements et extensions.'}, status=400)

        numero = request.data.get('numero_autorisation', '').strip()
        if not numero:
            return Response({'detail': 'Numéro d\'autorisation requis.'}, status=400)

        try:
            auto = Auto.objects.get(numero_auto=numero)
        except Auto.DoesNotExist:
            return Response({'detail': f'Autorisation {numero} introuvable.'}, status=404)

        # Lier à la demande
        demande.autorisation_parent = auto
        demande.save()

        # Lier au RenouvellementDetail si existe
        if hasattr(demande, 'renouvellement_detail'):
            rd = demande.renouvellement_detail
            rd.autorisation_origine = auto
            rd.numero_autorisation_original = numero
            rd.save()

        # Lier à l'ExtensionDetail si existe
        if hasattr(demande, 'extension_detail'):
            ed = demande.extension_detail
            ed.autorisation_origine = auto
            ed.numero_autorisation_original = numero
            ed.save()

        return Response({
            'message': f'Autorisation {numero} liée à la demande {demande.numero_ref}.',
            'autorisation': {
                'id': auto.id,
                'numero_auto': auto.numero_auto,
                'type': auto.type,
                'date_delivrance': str(auto.date_delivrance),
            }
        })